import pandas as pd
import os
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers


def export_table(df: pd.DataFrame, metadata: dict[str: str], format_rules: dict[str: str], filename: str):
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Data', startrow=0, startcol=2, index=False)  # колонки 0-1 для метадаты
        ws = writer.sheets['Data']
        metadata_cells = dict()

        # prepare metadata cells
        for i, (key, val) in enumerate(metadata.items(), 1):
            metadata_cells[f"A{i}"] = str(key)
            metadata_cells[f"B{i}"] = str(val)

        for cell, value in metadata_cells.items():
            ws[cell] = value

        # 3. Применяем форматирование: ОДИН РАЗ на весь столбец
        for idx, col_name in enumerate(df.columns):
            col_idx = idx + 3  # C = 3
            col_letter = get_column_letter(col_idx)

            if col_name in format_rules:
                fmt = format_rules[col_name]
                # Применяем формат ко всему столбцу (от строки 7 и ниже)
                for row in range(1, ws.max_row + 1):
                    cell = ws[f"{col_letter}{row}"]
                    cell.number_format = fmt

        # Автоширина
        ws.column_dimensions[get_column_letter(1)].width = 15
        ws.column_dimensions[get_column_letter(2)].width = 25
        for i, column in enumerate(df.columns, start=3):
            col_letter = get_column_letter(i)
            max_len = max(len(str(cell.value)) for cell in ws[col_letter][0:] if cell.value is not None)
            ws.column_dimensions[col_letter].width =min(max_len + 2, 50)


def main():
    metadata = {
        "Test Name:": f"Pidor",
        "Timestamp:": "Hz",
        "Operator:": "John Doe",
        "Notes:": "Power balance test (-1 to +1)"
    }

    # --- Форматирование: словарь ---
    format_rules = {
        'Power 1': '0',
        'Power 2': '0',
        'Power': '0',
        'Power balance': '0.00%',
        'Time': 'hh:mm:ss',
        'Hui': '0'
    }

    df = pd.DataFrame({
        'Time': pd.date_range(start='2025-01-01', periods=10, freq='1s'),
        'Power 1': [10, 20, 30, 40, 50, 60, 70, 80, 90, 100],
        'Power 2': [90, 80, 70, 60, 50, 40, 30, 20, 10, 0],
        'Power': [100] * 10,
        'Power balance': [-0.8, -0.6, -0.4, -0.2, 0, 0.2, 0.4, 0.6, 0.8, 1.0]
    })
    folder_name = "exports"
    name = "test_vehicle"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")  # todo: move to parameters
    raw_data_file = os.path.join(folder_name, f"test {name}_{timestamp}.xlsx")  # todo: change
    os.makedirs(folder_name, exist_ok=True)
    # export_table(df, metadata, format_rules, raw_data_file)
    export_table(df, {}, {}, raw_data_file)


if __name__ == "__main__":
    main()
